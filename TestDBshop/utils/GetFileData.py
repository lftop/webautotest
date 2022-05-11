from openpyxl import load_workbook
from common.common_path import *
import yaml


class ExcelCase:
    ID, MODULE, TEST_NAME, PREV_CONDITION, TESTDATA, TEST_STEP, TEST_EXPECTED = \
        ["用例编号", "测试模块", "用例名称", "前置条件", "测试数据", "步骤描述", "预期结果"]

    def __init__(self, **kwargs):
        self.__filename = kwargs.get("filename")
        self.__sheetname = kwargs.get("sheetname")

    @property
    def __load_excel(self):
        return load_workbook(filename=os.path.join(test_case_dir, self.__filename))[self.__sheetname]

    def get_max_row(self):
        return len(list(self.__load_excel.iter_rows()))

    def get_max_col(self):
        return len(list(self.__load_excel.iter_cols()))

    def get_items(self, S_Index=2, E_Index=0):
        data = []
        for rows in self.__load_excel.iter_rows(S_Index, E_Index if E_Index != 0 else self.get_max_row() - 1):
            result = dict()
            result.setdefault(self.ID, rows[0].value)
            result.setdefault(self.MODULE, rows[1].value)
            result.setdefault(self.TEST_NAME, rows[2].value)
            result.setdefault(self.PREV_CONDITION, rows[3].value)
            result.setdefault(self.TESTDATA, rows[4].value)
            result.setdefault(self.TEST_STEP, rows[5].value)
            result.setdefault(self.TEST_EXPECTED, rows[6].value)
            data.append(result)
        return data


class YamlData:
    def __init__(self, filename=""):
        self.__filename = filename

    def load_yaml(self):
        f = open(file=os.path.join(test_case_data_dir, self.__filename), encoding='utf-8', mode='r')
        return list(yaml.full_load_all(f.read()))


# if __name__ == "__main__":
#     # e = ExcelCase(filename="DBTESTCASE.xlsx", sheetname="DBSHOPADMIN")
#     #
#     # print(e.get_items())
#     y = YamlData("usermoduledata.yaml")
#     for row in y.load_yaml():
#         print(row)
